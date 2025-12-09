import pandas as pd
import re
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.db import transaction

# Импорт моделей
from registry.models import (
    ChemicalElement, RegistryConfig,
    Sec1Identification, Sec2Physical, Sec3ToxSanPin, Sec4ToxAir,
    Sec5ToxAcute, Sec6ToxRisks, Sec8EcoTox, Sec9Soil, Sec10Water,
    Sec11HazardClass, Sec12GHSClass, Sec13GHSLabel,
    Sec14Safety, Sec15Storage, Sec16Waste, Sec17Incidents,
    Sec18InternationalReg, Sec20Docs, Sec21Companies,
    Sec22Volumes, Sec23Extra
)

from .structures import SECTION_MAP

def get_field_info(model, field_name):
    try:
        return model._meta.get_field(field_name)
    except:
        return None

# =========================================================================
# 1. ГЕНЕРАЦИЯ ШАБЛОНА
# =========================================================================
def generate_excel_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Форма реестра"
    ws.freeze_panes = "A2"

    config_obj = RegistryConfig.objects.first()
    required_in_db = config_obj.required_fields if config_obj else []
    visible_in_db = config_obj.template_fields if config_obj else []

    header_font = Font(bold=True, color="FFFFFF", size=10)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'))

    current_col = 1

    for section_name, hex_color, model_cls, fields in SECTION_MAP:
        # Проверяем, есть ли в секции хоть одно поле, которое нужно показать
        section_has_visible_fields = False
        for _, db_f, is_sys in fields:
            if is_sys or (not visible_in_db) or (db_f in visible_in_db):
                section_has_visible_fields = True
                break

        if not section_has_visible_fields:
            continue

        fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

        for excel_header, db_field, is_system_required in fields:
            # Фильтрация по настройкам видимости (если настроены)
            if not is_system_required and visible_in_db and (db_field not in visible_in_db):
                continue

            is_mandatory = is_system_required or (db_field in required_in_db)
            final_header = excel_header + (" *" if is_mandatory else "")

            cell = ws.cell(row=1, column=current_col, value=final_header)
            cell.font = header_font
            cell.fill = fill
            cell.alignment = center_align
            cell.border = border

            col_letter = get_column_letter(current_col)
            ws.column_dimensions[col_letter].width = 25

            # Выпадающие списки (Data Validation)
            field_obj = get_field_info(model_cls, db_field)
            if field_obj and field_obj.choices:
                # Собираем варианты выбора
                options = [str(c[1]).strip() for c in field_obj.choices]
                # Excel имеет лимит 255 символов на формулу списка внутри ячейки
                formula_str = ",".join(options).replace("\"", "").replace("'", "")

                if len(formula_str) < 255:
                    dv = DataValidation(type="list", formula1=f'"{formula_str}"', allow_blank=True)
                    ws.add_data_validation(dv)
                    dv.add(f"{col_letter}2:{col_letter}2000")

            current_col += 1

    return wb

# =========================================================================
# 2. ИМПОРТ
# =========================================================================
def parse_boolean(value):
    v = str(value).lower().strip()
    return v in ['+', '1', 'yes', 'да', 'true', 'есть', 'y']

def process_excel_import(file, user):
    try:
        df = pd.read_excel(file)

        # === УЛУЧШЕННАЯ ОЧИСТКА ЗАГОЛОВКОВ ===
        clean_cols = []
        for col in df.columns:
            # 1. В строку
            c = str(col).strip()
            # 2. Убираем звездочку в конце (индикатор обязательности)
            if c.endswith("*"):
                c = c.replace("*", "").strip()
            # 3. Нормализуем пробелы (два пробела -> один)
            c = re.sub(r'\s+', ' ', c)
            clean_cols.append(c)

        df.columns = clean_cols
        # Заменяем NaN на пустую строку, чтобы не падать
        df = df.fillna("")

    except Exception as e:
        return {"success": 0, "errors": [f"Ошибка чтения файла (структура Excel): {str(e)}"]}

    report = {"success": 0, "errors": []}

    # 1. Создаем карты маппинга
    FLAT_MAP = {}      # Заголовок Excel (чистый) -> (КлассМодели, ИмяПоляБД)
    HUMAN_NAMES = {}   # ИмяПоляБД -> Заголовок Excel (для красивых ошибок)

    for _, _, model, fields in SECTION_MAP:
        for ex_head, db_field, _ in fields:
            # Нормализуем и ключ маппинга тоже, чтобы совпадало наверняка
            clean_head = re.sub(r'\s+', ' ', ex_head.strip())
            FLAT_MAP[clean_head] = (model, db_field)
            HUMAN_NAMES[db_field] = ex_head

    # 2. Итерация по строкам
    for index, row in df.iterrows():
        row_num = index + 2 # В Excel строки начинаются с 1, плюс заголовок
        try:
            with transaction.atomic():
                extracted_data = {}

                # A. Парсинг данных
                for col_name in df.columns:
                    col_name_clean = str(col_name).strip()

                    if col_name_clean not in FLAT_MAP:
                        continue # Колонка не из нашего шаблона

                    target_model, target_field = FLAT_MAP[col_name_clean]
                    raw_value = str(row[col_name]).strip()

                    if not raw_value:
                        continue

                    field_obj = get_field_info(target_model, target_field)
                    final_value = raw_value

                    # Если поле с выбором (choices), пытаемся найти ключ по значению
                    # (Пользователь видит "Твердое вещество", а в базу надо писать "SOLID")
                    if field_obj and field_obj.choices:
                        for k, label in field_obj.choices:
                            if raw_value.lower() == str(label).lower() or raw_value.lower() == str(k).lower():
                                final_value = k
                                break

                    # Если булево поле
                    if field_obj and field_obj.get_internal_type() == 'BooleanField':
                        final_value = parse_boolean(raw_value)

                    # Собираем данные по моделям
                    if target_model not in extracted_data:
                        extracted_data[target_model] = {}
                    extracted_data[target_model][target_field] = final_value

                # Б. ВАЛИДАЦИЯ (Динамическая)
                config_obj = RegistryConfig.objects.first()
                required_list = config_obj.required_fields if config_obj else []

                if required_list:
                    for field_name in required_list:
                        # Ищем это поле среди всех собранных данных
                        found_val = False
                        for model_class, fields_dict in extracted_data.items():
                            if field_name in fields_dict:
                                val = fields_dict[field_name]
                                if val is not None and val != "":
                                    found_val = True
                                    break

                        # Если не нашли значение для обязательного поля
                        if not found_val:
                            human_name = HUMAN_NAMES.get(field_name, field_name)
                            raise ValueError(f"Поле '{human_name}' обязательно для заполнения.")

                # В. Системная валидация (Главное название)
                if ChemicalElement not in extracted_data or 'primary_name_ru' not in extracted_data[ChemicalElement]:
                     # Если вся строка пустая - пропускаем молча
                     if not row.empty and any(str(x).strip() for x in row):
                        raise ValueError("Отсутствует 'Название вещества (RU)'")
                     continue

                # Г. Сохранение в БД
                # Сначала главную модель
                element = ChemicalElement.objects.create(
                    created_by=user,
                    status=ChemicalElement.Status.DRAFT,
                    **extracted_data[ChemicalElement]
                )

                # Потом все секции
                for model_cls, fields_dict in extracted_data.items():
                    if model_cls == ChemicalElement: continue
                    model_cls.objects.create(element=element, **fields_dict)

                # Создаем пустые записи для недостающих секций (чтобы админка не падала)
                for rel_obj in ChemicalElement._meta.related_objects:
                    if rel_obj.one_to_one and rel_obj.name.startswith('sec'):
                        if not hasattr(element, rel_obj.name):
                            rel_obj.related_model.objects.create(element=element)

                report["success"] += 1

        except Exception as e:
            report["errors"].append(f"Строка {row_num}: {str(e)}")

    return report